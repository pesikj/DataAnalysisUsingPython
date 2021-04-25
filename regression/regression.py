import numpy as np
import statsmodels.api as sm
from openpyxl import load_workbook

MODE = "SP"

x = []
y = []
predAlpha = 0.1
coefAlpha = 0.1

if MODE == "SP":
    wb = load_workbook(filename='P_2SM_zadani-1.xlsx')
    ws = wb.worksheets[1]
    student_number = input("Zadej osobní číslo: ")

    for row in range(1, 150):
        cell = ws.cell(column=1, row=row)
        if cell.value == "číslo studenta":
            for col in range(4, 50):
                cell = ws.cell(column=col, row=row)
                if cell.value:
                    x.append(cell.value)
        if cell.value == student_number:
            for col in range(4, 50):
                x0 = ws.cell(column=3, row=row).value
                cell = ws.cell(column=col, row=row)
                if cell.value:
                    y.append(ws.cell(column=col, row=row).value)

else:
    wb = load_workbook(filename = 'data.xlsx')
    ws = wb.active
    yCol = 6
    xCol = 2
    x0 = 40

    for row in range(1, 26):
        cell = ws.cell(column=yCol, row=row).value
        y.append(ws.cell(column=yCol, row=row).value)
        x.append(ws.cell(column=xCol, row=row).value)

x = sm.add_constant(x)

mod = sm.OLS(y, x)
res = mod.fit()
print(f"Parameters are: b0 = {round(res.params[0], 4)}, b1 = {round(res.params[1], 4)}")
print(f"R2 = {round(res.rsquared, 4)}")
confIntervals = res.conf_int(alpha=coefAlpha)
print(f"Conf Intervals ({(1 - coefAlpha) * 100 } %), Beta0: < {round(confIntervals[0][0], 4)}, {round(confIntervals[0][1], 4)} >, Beta1: < {round(confIntervals[1][0], 4)}, {round(confIntervals[1][1], 4)} >")

predAvarege = res.get_prediction(exog=[1, x0]).conf_int(alpha=predAlpha)
print(f"Average Conf Intervals for x = {x0} ({(1 - predAlpha) * 100 } %): < {round(predAvarege[0][0], 4)}, {round(predAvarege[0][1], 4)} > ")
predObserved = res.get_prediction(exog=[1, x0]).conf_int(obs=True, alpha=predAlpha)
print(f"Observed Conf Intervals for x = {x0} ({(1 - predAlpha) * 100 } %): < {round(predObserved[0][0], 4)}, {round(predObserved[0][1], 4)} > ")