import numpy as np
from openpyxl import load_workbook
from scipy import stats
from scipy.stats import t, f, norm

MODE = "SP"


def get_data_for_region(regionLetter):
    col = ord(regionLetter) - 63
    data = []
    for row in range(2, 28):
        cell = ws.cell(column=col, row=row)
        if cell.value:
            data.append(cell.value)
    return np.array(data).astype(np.float)


def get_data_for_student(student_number):
    col = 1
    data1 = []
    data2 = []
    row = 2
    while True:
        cell = ws.cell(column=col, row=row)
        if cell.value == student_number:
            for col in range(3, 20):
                cell = ws.cell(column=col, row=row)
                if cell.value:
                    data1.append(cell.value)
            row += 1
            for col in range(3, 20):
                cell = ws.cell(column=col, row=row)
                if cell.value:
                    data2.append(cell.value)
            break
        row += 1
    return data1, data2


def perform_ks_test(data, alpha=0.05):
    normed_data = []
    for value in data:
        normed_data.append((value - 61)/9)
    statistics, pvalue = stats.kstest(normed_data, "norm")
    print(f"Test statistics: {round(statistics, 4)}")
    if pvalue < alpha:
        print("H0 rejected")
    else:
        print("H0 NOT rejected")


def perform_f_test(data1, data2, alpha, alternative):
    statistics = np.var(data1, ddof=1) / np.var(data2, ddof=1)
    df1 = len(data1) - 1
    df2 = len(data2) - 1
    if alternative == "≠":
        print(f"< 0, {round(f.ppf(alpha / 2, df1, df2), 3)} > < {round(f.ppf(1 - alpha / 2, df1, df2), 3)} , ∞)")
    elif alternative == "<":
        print(f"< 0, {round(f.ppf(alpha, df1, df2), 3)} >")
    elif alternative == ">":
        print(f"< {round(f.ppf(1 - alpha, df1, df2), 3)} , ∞)")
    else:
        print("Incorrect alternative")
        return
    print(f"Test statistics: {round(statistics, 4)}")
    median = f.ppf(0.5, df1, df2)
    if alternative == "≠":
        if statistics < median:
            pvalue = f.cdf(statistics, df1, df2) * 2
        else:
            pvalue = (1 - f.cdf(statistics, df1, df2)) * 2
    elif (alternative == "<" and statistics < median) or (alternative == ">" and statistics > median):
        pvalue = f.cdf(statistics, df1, df2)
    else:
        pvalue = 1 - f.cdf(statistics, df1, df2)
    
    if pvalue < alpha:
        print("H0 rejected")
    else:
        print("H0 NOT rejected")
    print(f"p-value: {round(pvalue, 4)}")
    return pvalue


def perform_t_test(data1, data2, equalVar, alpha, alternative):
    equalVar = equalVar.upper() == "Y"
    df = len(data1) + len(data2) - 2
    if alternative == "≠":
        print(f"( - ∞, {round(t.ppf(alpha / 2, df), 3)} > < {round(t.ppf(1 - alpha / 2, df), 3)} , ∞)")
    elif alternative == "<":
        print(f"( - ∞, {round(t.ppf(alpha, df), 3)} >")
    elif alternative == ">":
        print(f"< {round(t.ppf(1 - alpha, df), 3)} , ∞)")
    else:
        print("Incorrect alternative")
        return
    results = stats.ttest_ind(data1, data2, equal_var=equalVar)
    statistics = results[0]
    print(f"Test statistics: {round(statistics, 4)}")
    if alternative == "≠":
        pvalue = results[1]
    elif (alternative == "<" and statistics < 0) or (alternative == ">" and statistics > 0):
        pvalue = results[1] / 2
    else:
        pvalue = 1 - results[1] / 2
    if pvalue < alpha:
        print("H0 rejected")
    else:
        print("H0 NOT rejected")
    print(f"p-value: {round(pvalue, 4)}")


if MODE == "SP":
    wb = load_workbook(filename='P_2SM_zadani-1.xlsx')
    ws = wb.worksheets[0]
    osobni_cislo = input("Zadej osobní číslo: ")
    data1, data2 = get_data_for_student(osobni_cislo)
    perform_ks_test(data1)
    pval = perform_f_test(data1, data2, 0.05, "≠")
    perform_t_test(data1, data2, "y" if pval > 0.05 else "n", 0.05, "≠")
else:
    wb = load_workbook(filename='data.xlsx')
    ws = wb.active
    testType = input("Select test type [f/t/median]: ")
    testType = testType.upper()
    if testType in ('T', 'F'):
        firstLocality = input("First locality [A-L]: ")
        secondLocality = input("First locality [A-L]: ")
        if testType == 'T':
            equalVar = input("Equal Variances [y/n]: ")
        alpha = input("Alpha: ")
        alpha = float(alpha)
        alternative = input("Alternative [neq/</>]: ")
        if alternative == "neq":
            alternative = "≠"
        data1 = get_data_for_region(firstLocality.upper())
        data2 = get_data_for_region(secondLocality.upper())

        if testType == 'T':
            print(f"H0: μ {firstLocality} = μ {secondLocality}")
            print(f"H1: μ {firstLocality} {alternative} μ {secondLocality}")
            perform_t_test(data1, data2, equalVar, alpha, alternative)
        elif testType == 'F':
            print(f"H0: σ2 {firstLocality} = σ2 {secondLocality}")
            print(f"H1: σ2 {firstLocality} {alternative} σ2 {secondLocality}")
            perform_f_test(data1, data2, alpha, alternative)

